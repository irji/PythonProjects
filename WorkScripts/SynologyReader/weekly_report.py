from synology_drive_api.drive import SynologyDrive
from openpyxl import load_workbook, Workbook
import pandas as pd
import datetime

exclude_headers = {"Заполнение отчета",
                   "Переписка в почте, сообщения в редмайне, рокетчате, звонки в зуме. Обсуждения с коллегами рабочих вопросов.",
                   "Ежедневное собрание группы",
                   "Ежедневное собрание группы + подгруппы",
                   "Интерфейс",
                   "Еженедельная лекция",
                   "Корпоратив"}

user_name = "georgii.kostin"
employee_name = "Костин Георгий"
user_password = "5MGc/8cac"
server_path = "synoffice.local.rfdyn.ru"
template_name = "Kostin_Georgii_template.xlsx"

def get_monday (date):
  return date  - datetime.timedelta(days=date.weekday() % 7)

def get_friday (date):
  return get_monday (date) + datetime.timedelta (days = 4)



def aggregate_comments(input_dataframe):
    input_dataframe = input_dataframe.reset_index(drop=True).fillna("")

    # делаем группировку по колонке с темой, чтобы одинаковые задачи собрать
    grouped_dataframe = input_dataframe.groupby(['Тема'])
    result_dataframe = None

    for header, df in grouped_dataframe:
        new_comment = ""
        new_id = None
        new_header = str(header[0]).rstrip()

        # проверм что темы нет в списке исключений и объединяем все комментарии
        if new_header not in exclude_headers:
            for row in df.iterrows():
                comment = str(row[1]["Комментарии"])
                new_id = row[1]["#"]
                #new_header = row[1]["Тема"]

                if comment not in new_comment:
                    new_comment = new_comment + comment + "; "

            new_comment = new_comment.rstrip().removesuffix(";")
            data = {"#":[new_id], "Тема":[new_header], "Комментарии":[new_comment]}

            result_dataframe = pd.concat([result_dataframe, pd.DataFrame(data)], axis=0)

    return result_dataframe



def create_path_list(folder_name1, folder_name2):
    paths = []

    paths.append("/team-folders/Reports/SUPPORT/" + folder_name1 + "/3. Report Support Среда.osheet")
    paths.append("/team-folders/Reports/SUPPORT/" + folder_name1 + "/4. Report Support Четверг.osheet")
    paths.append("/team-folders/Reports/SUPPORT/" + folder_name1 + "/5. Report Support Пятница.osheet")

    paths.append("/team-folders/Reports/SUPPORT/" + folder_name2 + "/1. Report Support Понедельник.osheet")
    paths.append("/team-folders/Reports/SUPPORT/" + folder_name2 + "/2. Report Support Вторник.osheet")

    return  paths

def create_archive_path_list(begin_date :datetime.date):
    paths = []
    new_date = begin_date

    with SynologyDrive(user_name, user_password, server_path, dsm_version="7") as synology_drive:
        folders = synology_drive.list_folder("/team-folders/Reports/SUPPORT/Archive")["data"]["items"]
        folder_list = []

        for fldr in folders:
            folder_list.append(fldr.get("name"))

        while new_date < datetime.datetime.today().date():
            folder_path = ""

            new_month_id = "0" + str(new_date.month) if new_date.month < 10 else str(new_date.month)
            new_day_id = "0" + str(new_date.day) if new_date.day < 10 else str(new_date.day)
            new_folder_name = "Week " + str(new_date.year) + new_month_id + new_day_id

            if new_folder_name in folder_list:
                folder_path = "/team-folders/Reports/SUPPORT/Archive/" + new_folder_name
            else:
                folder_path = "/team-folders/Reports/SUPPORT/" + new_folder_name

            paths.append(folder_path + "/1. Report Support Понедельник.osheet")
            paths.append(folder_path + "/3. Report Support Среда.osheet")
            paths.append(folder_path + "/4. Report Support Четверг.osheet")
            paths.append(folder_path + "/2. Report Support Вторник.osheet")
            paths.append(folder_path + "/5. Report Support Пятница.osheet")

            new_date = new_date + datetime.timedelta(7)

    return  paths



def get_task_list_for_weekly_report():
    vertical_concat = None

    with SynologyDrive(user_name, user_password, server_path, dsm_version="7") as synology_drive:
        folders = synology_drive.list_folder("/team-folders/Reports/SUPPORT")["data"]["items"]

        path_list = create_path_list(folders[1]["name"], folders[2]["name"])

        for path in path_list:
            #if item["name"] != "Archive" and item["name"] != "Week Шаблон":
                #print(item["name"])
                #path = "/team-folders/Reports/SUPPORT/" + item["name"] + "/3. Report Support Среда.osheet"

                report_file = synology_drive.download_synology_office_file(path)
                tabl = pd.read_excel(report_file, sheet_name=None, na_values="nan")["Support"]

                # Фильтруем dataframe только для себя + удаляем строки с пустым значением в колонке 'Тема'
                filtered_tabl = tabl[tabl["Сотрудник"] == employee_name].dropna(subset=["Тема"])
                vertical_concat = pd.concat([vertical_concat, filtered_tabl], axis=0)

    vertical_concat.sort_values(by="Тема", inplace=True)
    task_list = aggregate_comments(vertical_concat)
    idx = pd.Index(range(0, len(task_list.index), 1))
    task_list = task_list.set_index(idx)

    return task_list

def get_task_list_for_quarter_report(begin_date :datetime.date):
    vertical_concat = None
    date1 = get_monday(begin_date)
    paths = create_archive_path_list(date1)

    with SynologyDrive(user_name, user_password, server_path, dsm_version="7") as synology_drive:
        for path in paths:
                report_file = synology_drive.download_synology_office_file(path)
                tabl = pd.read_excel(report_file, sheet_name=None, na_values="nan")["Support"]

                # Фильтруем dataframe только для себя + удаляем строки с пустым значением в колонке 'Тема'
                filtered_tabl = tabl[tabl["Сотрудник"] == employee_name].dropna(subset=["Тема"])
                vertical_concat = pd.concat([vertical_concat, filtered_tabl], axis=0)

    vertical_concat.sort_values(by="Тема", inplace=True)
    task_list = aggregate_comments(vertical_concat)
    idx = pd.Index(range(0, len(task_list.index), 1))
    task_list = task_list.set_index(idx)

    return task_list



def fill_weekly_xls_report(task_list):
    date_string = ""
    new_month_id = "0" + str(datetime.datetime.today().month) if datetime.datetime.today().month < 10 else str(datetime.datetime.today().month)
    new_day_id = "0" + str(datetime.datetime.today().day) if datetime.datetime.today().day < 10 else str(datetime.datetime.today().day)

    date_string = str(datetime.datetime.today().year) + new_month_id + new_day_id
    new_file_name = template_name.replace("template", date_string)
    #shutil.copy(template_name, new_file_name)

    wb = load_workbook(filename = template_name)
    sheet_ranges = wb['Tasks']

    for i,task in task_list.iterrows():
        sheet_ranges.cell(row=i + 3, column=1).value = task["#"]
        sheet_ranges.cell(row=i + 3, column=4).value = task["Тема"]
        sheet_ranges.cell(row=i + 3, column=5).value = task["Комментарии"]

    wb.save(new_file_name)

def fill_quarter_xls_report(task_list):
    wb = Workbook()
    sheet_ranges = wb.active

    for i,task in task_list.iterrows():
        sheet_ranges.cell(row=i+1, column=1).value = task["#"]
        sheet_ranges.cell(row=i+1, column=2).value = task["Тема"]
        sheet_ranges.cell(row=i+1, column=3).value = task["Комментарии"]

    wb.save("Report.xlsx")

if __name__ == '__main__':
    #task_list = get_task_list_for_weekly_report()
    #fill_xls_report(task_list)

    task_list = get_task_list_for_quarter_report(datetime.date(2025, 5, 28))
    fill_quarter_xls_report(task_list)