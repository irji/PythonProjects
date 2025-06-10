from openpyxl import load_workbook
import os

user_name = "Костин Георгий"
input_dir = "D:/Отчеты/"

print("File|Name|Date|Id|Description|Comment")

dirs = os.listdir(input_dir)

for sub_dirs in dirs:
    files = os.listdir(input_dir + sub_dirs)

    for file in files:
        #print(file)
        #file = "D:/Отчеты/Week 20240826/1. Report Support Понедельник.xlsx"

        xl_workbook = load_workbook(filename=input_dir + sub_dirs + "/" + file)
        sheet_ranges = xl_workbook["Support"]

        for row in range(1,500):
            if sheet_ranges.cell(row=row, column=3).value == user_name:
                if sheet_ranges.cell(row=row, column=5).value is not None:
                    rep_day = str(sheet_ranges.cell(row=row, column=2).value).strip()
                    rep_id = str(sheet_ranges.cell(row=row, column=4).value).strip()
                    rep_description = str(sheet_ranges.cell(row=row, column=5).value).strip()
                    rep_comment = str(sheet_ranges.cell(row=row, column=10).value).strip()

                    print("{}|{}|{}|{}|{}|{}".format(sub_dirs, file, rep_day, rep_id, rep_description, rep_comment))